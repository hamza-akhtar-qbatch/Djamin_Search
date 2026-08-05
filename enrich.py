#!/usr/bin/env python3
"""
Djaminn audio enrichment pipeline.

For each selected song: download MIXDOWN_AUDIO -> Gemini audio analysis
(full transcription + feel + instrumentation + musical facts) -> embedding
of the composed profile -> checkpointed JSONL. Finally compiles the
web-ready index consumed by the Next.js app.

Usage:
    python3 enrich.py --limit 150            # select + enrich subset
    python3 enrich.py --limit 150 --workers 4
    python3 enrich.py --compile              # just rebuild web/data/index.json

Resume-safe: songs already in enriched.jsonl are skipped on re-run.
Stdlib + openpyxl only.
"""

import argparse
import base64
import json
import math
import os
import re
import sys
import time
import unicodedata
import urllib.request
import urllib.error
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import defaultdict, Counter
from datetime import datetime

import openpyxl

XLSX = "DJaminn_Top_Artists_Songs_original_candidate_data.xlsx"
AUDIO_DIR = "audio_cache"
CHECKPOINT = "enriched.jsonl"
OUT_INDEX = "web/data/index.json"

# Free-tier daily quotas are per-model; the ladder falls through to the next
# audio-capable model when one bucket runs dry.
MODEL_LADDER = [m for m in (
    os.environ.get("GEN_MODEL"),
    "gemini-3.5-flash",
    "gemini-3-flash-preview",
    "gemini-3.1-flash-lite",
) if m]
EMBED_MODEL = "gemini-embedding-001"
EMBED_DIMS = 768
API = "https://generativelanguage.googleapis.com/v1beta/models"

HOUSE_ACCOUNTS = {"Djaminn Artists"}

ANALYSIS_SCHEMA = {
    "type": "object",
    "properties": {
        "description": {"type": "string", "description": "3-4 sentences describing the sonic experience of the track for a music discovery index"},
        "transcription": {"type": "string", "description": "Verbatim lyrics transcription in the original language. Empty string if instrumental or unintelligible."},
        "lyrics_gist": {"type": "string", "description": "1-2 sentence English summary of what the lyrics are about. Empty if instrumental."},
        "instruments": {"type": "array", "items": {"type": "string"}, "description": "Instruments heard, most prominent first"},
        "moods": {"type": "array", "items": {"type": "string"}, "description": "Moods/feelings/emotions the track evokes, 3-6 lowercase words"},
        "genres": {"type": "array", "items": {"type": "string"}, "description": "Genre and subgenres heard, lowercase"},
        "energy": {"type": "string", "enum": ["low", "medium", "high"]},
        "has_vocals": {"type": "boolean"},
        "vocal_type": {"type": "string", "enum": ["none", "male", "female", "group", "unclear"]},
        "vocal_delivery": {"type": "string", "enum": ["none", "sung", "rapped", "spoken", "mixed"]},
        "language": {"type": "string", "description": "ISO 639-1 code of lyrics language, or 'none' if instrumental"},
        "bpm_estimate": {"type": "integer", "description": "Estimated tempo in BPM"},
        "tempo_feel": {"type": "string", "enum": ["slow", "mid", "fast"]},
        "key_guess": {"type": "string", "description": "Estimated musical key e.g. 'A minor', or 'unclear'"},
        "production": {"type": "string", "description": "Production character: lo-fi/polished, live/studio, loop or full arrangement, notable texture. 1 sentence."},
        "is_cover_likely": {"type": "boolean", "description": "True if this sounds like a cover of a known song"},
        "good_for_remixing": {"type": "boolean", "description": "True if instrumental/loop/beat-like material a producer could build on"},
        "use_cases": {"type": "array", "items": {"type": "string"}, "description": "Listening/use contexts e.g. study, workout, late night drive, remix material, jam backing track"},
        "sounds_like": {"type": "array", "items": {"type": "string"}, "description": "1-3 well-known artists this track's sound resembles"},
    },
    "required": ["description", "transcription", "lyrics_gist", "instruments", "moods",
                 "genres", "energy", "has_vocals", "vocal_type", "vocal_delivery",
                 "language", "bpm_estimate", "tempo_feel", "key_guess", "production",
                 "is_cover_likely", "good_for_remixing", "use_cases", "sounds_like"],
}

PROMPT = """You are indexing music for a discovery/search platform for musicians.
Listen to this track and extract a complete, factual profile of it.
Transcribe any lyrics verbatim in their original language (Portuguese, Spanish, English, Thai and others all occur).
Keep the transcription to at most 300 words — for longer songs transcribe the first verses and chorus only.
Describe only what you actually hear — do not invent details. If something is unclear, say so.
Return JSON matching the schema."""

SHORT_RETRY_NOTE = "\nIMPORTANT: your previous output was truncated. Keep the transcription under 100 words this time."


def api_keys():
    names = ["GEMINI_API_KEY"] + [f"GEMINI_API_KEY_{i}" for i in range(2, 10)]
    keys = []
    for name in names:
        v = os.environ.get(name)
        if v and v not in keys:
            keys.append(v)
    if os.path.exists(".env"):
        with open(".env") as f:
            for line in f:
                for name in names:
                    if line.startswith(name + "="):
                        v = line.split("=", 1)[1].strip()
                        if v and v not in keys:
                            keys.append(v)
    if not keys:
        sys.exit("GEMINI_API_KEY not found in env or .env")
    return keys


ALL_KEYS = api_keys()
KEY = ALL_KEYS[0]


def post_json(url, body, tries=6):
    data = json.dumps(body).encode()
    for attempt in range(tries):
        try:
            req = urllib.request.Request(url, data=data,
                                         headers={"Content-Type": "application/json"})
            with urllib.request.urlopen(req, timeout=300) as r:
                return json.load(r)
        except urllib.error.HTTPError as e:
            if e.code == 429:
                # quota on this key -> roll to the next key before backing off
                m = re.search(r"key=([^&]+)", url)
                cur = m.group(1) if m else None
                if cur in ALL_KEYS and ALL_KEYS.index(cur) + 1 < len(ALL_KEYS):
                    nxt = ALL_KEYS[ALL_KEYS.index(cur) + 1]
                    url = url.replace(f"key={cur}", f"key={nxt}")
                    print("    quota hit -> switching to fallback API key", flush=True)
                    continue
            if e.code in (429, 500, 502, 503) and attempt < tries - 1:
                retry = e.headers.get("Retry-After")
                wait = float(retry) if retry else min(60, 2 ** attempt * 5)
                print(f"    HTTP {e.code}, retrying in {wait:.0f}s", flush=True)
                time.sleep(wait)
                continue
            raise RuntimeError(f"HTTP {e.code}: {e.read()[:300]}") from e
        except (urllib.error.URLError, TimeoutError) as e:
            if attempt < tries - 1:
                time.sleep(min(60, 2 ** attempt * 5))
                continue
            raise


_dead_models = set()
_model_strikes = {}
_ladder_lock = __import__("threading").Lock()


def analyze_audio(mp3_path):
    audio = base64.b64encode(open(mp3_path, "rb").read()).decode()
    body = {
        "contents": [{"parts": [
            {"inline_data": {"mime_type": "audio/mpeg", "data": audio}},
            {"text": PROMPT},
        ]}],
        "generationConfig": {
            "responseMimeType": "application/json",
            "responseJsonSchema": ANALYSIS_SCHEMA,
            "temperature": 0,
        },
    }
    last_err = None
    for model in MODEL_LADDER:
        with _ladder_lock:
            if model in _dead_models:
                continue
        try:
            r = post_json(f"{API}/{model}:generateContent?key={KEY}", body, tries=4)
            with _ladder_lock:
                _model_strikes[model] = 0
            raw = r["candidates"][0]["content"]["parts"][0]["text"]
            try:
                return json.loads(raw)
            except json.JSONDecodeError:
                # long lyrics can truncate the JSON — retry once, shorter
                retry_body = json.loads(json.dumps(body))
                retry_body["contents"][0]["parts"][1]["text"] = PROMPT + SHORT_RETRY_NOTE
                r = post_json(f"{API}/{model}:generateContent?key={KEY}", retry_body, tries=3)
                return json.loads(r["candidates"][0]["content"]["parts"][0]["text"])
        except RuntimeError as e:
            last_err = e
            if "429" in str(e):
                with _ladder_lock:
                    _model_strikes[model] = _model_strikes.get(model, 0) + 1
                    if _model_strikes[model] >= 2:
                        _dead_models.add(model)
                        print(f"    model {model} quota exhausted -> falling through",
                              flush=True)
            else:
                raise
    raise last_err or RuntimeError("all models exhausted")


def embed(text, task="RETRIEVAL_DOCUMENT"):
    body = {"content": {"parts": [{"text": text[:8000]}]},
            "taskType": task, "outputDimensionality": EMBED_DIMS}
    r = post_json(f"{API}/{EMBED_MODEL}:embedContent?key={KEY}", body)
    return r["embedding"]["values"]


def profile_text(song, analysis):
    """Composed document that gets embedded — the searchable meaning."""
    a = analysis
    parts = [
        f"{song['SONG_TITLE']} by {song['ARTIST_NAME']}.",
        a["description"],
        "Instruments: " + ", ".join(a["instruments"]) + ".",
        "Mood: " + ", ".join(a["moods"]) + ".",
        "Genres: " + ", ".join(a["genres"]) + ".",
        f"Energy {a['energy']}, tempo {a['tempo_feel']} ({a['bpm_estimate']} bpm).",
        a["production"],
    ]
    if a["has_vocals"]:
        parts.append(f"{a['vocal_type']} vocals, {a['vocal_delivery']}.")
        if a["lyrics_gist"]:
            parts.append("Lyrics about: " + a["lyrics_gist"])
        # English translation carries cross-lingual meaning; verbatim original
        # keeps exact-phrase search working
        if a.get("lyrics_translation_en"):
            parts.append("Lyrics (English): " + a["lyrics_translation_en"][:1200])
        if a["transcription"]:
            parts.append("Lyrics: " + a["transcription"][:800])
    else:
        parts.append("Instrumental, no vocals.")
    if a.get("themes"):
        parts.append("Themes: " + ", ".join(a["themes"]) + ".")
    if a.get("sentiment"):
        parts.append("Sentiment: " + a["sentiment"] + ".")
    if a["use_cases"]:
        parts.append("Good for: " + ", ".join(a["use_cases"]) + ".")
    if a["sounds_like"]:
        parts.append("Sounds like: " + ", ".join(a["sounds_like"]) + ".")
    if song.get("artist_genres"):
        parts.append("Artist genres: " + ", ".join(song["artist_genres"]) + ".")
    return " ".join(parts)


# ---------------------------------------------------------------- data


def load_songs():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    def sheet(name):
        rows = wb[name].iter_rows(values_only=True)
        hdr = next(rows)
        return [dict(zip(hdr, r)) for r in rows if any(c is not None for c in r)]

    users = {u["USER_ID"]: u for u in sheet("User")}
    songs = sheet("Songs")
    for s in songs:
        u = users.get(s["USER_ID"], {})
        s["artist_genres"] = [g.strip() for g in str(u.get("GENRES") or "").split(",") if g.strip()]
        s["region"] = u.get("REGION")
        s["artist_rating"] = u.get("RATING_SCORE")
        s["artist_followers"] = u.get("FOLLOWER_COUNT")
    return songs


def select_subset(songs, limit):
    """Top-3 per artist by playback for coverage, then fill by global playback."""
    by_artist = defaultdict(list)
    for s in songs:
        by_artist[s["USER_ID"]].append(s)
    picked, picked_ids = [], set()
    for ss in by_artist.values():
        for s in sorted(ss, key=lambda x: x["PLAYBACK_COUNT"] or 0, reverse=True)[:3]:
            picked.append(s)
            picked_ids.add(s["SONG_ID"])
    if len(picked) < limit:
        rest = sorted((s for s in songs if s["SONG_ID"] not in picked_ids),
                      key=lambda x: x["PLAYBACK_COUNT"] or 0, reverse=True)
        picked.extend(rest[:limit - len(picked)])
    return picked[:limit]


def download(song):
    os.makedirs(AUDIO_DIR, exist_ok=True)
    path = os.path.join(AUDIO_DIR, song["SONG_ID"] + ".mp3")
    if os.path.exists(path) and os.path.getsize(path) > 1000:
        return path
    url = song["MIXDOWN_AUDIO"] or song["AUDIO_URL"]
    urllib.request.urlretrieve(url, path)
    return path


# ---------------------------------------------------------------- pipeline


def done_ids():
    if not os.path.exists(CHECKPOINT):
        return set()
    with open(CHECKPOINT) as f:
        return {json.loads(line)["song_id"] for line in f if line.strip()}


def enrich_one(song):
    path = download(song)
    size_mb = os.path.getsize(path) / 1e6
    if size_mb > 19:
        return {"song_id": song["SONG_ID"], "error": f"file too large ({size_mb:.0f}MB)"}
    analysis = analyze_audio(path)
    text = profile_text(song, analysis)
    vector = embed(text)
    return {
        "song_id": song["SONG_ID"],
        "title": str(song["SONG_TITLE"]),
        "artist": str(song["ARTIST_NAME"]),
        "artist_id": song["USER_ID"],
        "audio_url": song["MIXDOWN_AUDIO"] or song["AUDIO_URL"],
        "upload_date": str(song["UPLOAD_DATE"]),
        "playback_count": song["PLAYBACK_COUNT"] or 0,
        "duration_seconds": song["DURATION_SECONDS"],
        "region": song["region"],
        "artist_genres": song["artist_genres"],
        "artist_rating": song["artist_rating"],
        "artist_followers": song["artist_followers"],
        "analysis": analysis,
        "profile_text": text,
        "embedding": vector,
    }


def run(limit, workers):
    songs = load_songs()
    subset = select_subset(songs, limit)
    skip = done_ids()
    todo = [s for s in subset if s["SONG_ID"] not in skip]
    print(f"Selected {len(subset)} songs; {len(skip)} already enriched; {len(todo)} to do.")

    ok = err = 0
    with open(CHECKPOINT, "a") as out, ThreadPoolExecutor(max_workers=workers) as ex:
        futures = {ex.submit(enrich_one, s): s for s in todo}
        for i, fut in enumerate(as_completed(futures), 1):
            s = futures[fut]
            try:
                rec = fut.result()
                if "error" in rec:
                    err += 1
                    print(f"[{i}/{len(todo)}] SKIP {s['SONG_TITLE'][:40]}: {rec['error']}", flush=True)
                else:
                    ok += 1
                    out.write(json.dumps(rec, ensure_ascii=False) + "\n")
                    out.flush()
                    a = rec["analysis"]
                    print(f"[{i}/{len(todo)}] OK {s['SONG_TITLE'][:40]!r} — "
                          f"{'/'.join(a['genres'][:2])}, {a['energy']} energy, "
                          f"vocals={a['vocal_type']}", flush=True)
            except Exception as e:
                err += 1
                print(f"[{i}/{len(todo)}] FAIL {s['SONG_TITLE'][:40]}: {e}", flush=True)
    print(f"\nDone: {ok} enriched, {err} failed. Checkpoint: {CHECKPOINT}")


def compile_index():
    if not os.path.exists(CHECKPOINT):
        sys.exit("No enriched.jsonl yet — run enrichment first.")
    records, seen = [], set()
    sources = [CHECKPOINT] + (["test_tracks.jsonl"] if os.path.exists("test_tracks.jsonl") else [])
    for src in sources:
        with open(src) as f:
            for line in f:
                if not line.strip():
                    continue
                try:
                    r = json.loads(line)
                except json.JSONDecodeError:
                    continue  # partial line from a concurrent writer
                if "error" in r or r["song_id"] in seen:
                    continue
                seen.add(r["song_id"])
                r["title"] = str(r.get("title", ""))
                r["artist"] = str(r.get("artist", ""))
                records.append(r)

    anchor = max(datetime.fromisoformat(r["upload_date"].replace("Z", "+00:00"))
                 for r in records)
    max_pb = max(r["playback_count"] for r in records) or 1
    for r in records:
        days = (anchor - datetime.fromisoformat(r["upload_date"].replace("Z", "+00:00"))).days
        pop = math.log1p(r["playback_count"]) / math.log1p(max_pb)
        rec = math.exp(-days / 120)
        r["prior"] = round(0.7 * pop + 0.3 * rec, 4)
        if r["artist"] in HOUSE_ACCOUNTS:
            r["prior"] = round(r["prior"] * 0.4, 4)

    artists = build_artists(records)
    os.makedirs(os.path.dirname(OUT_INDEX), exist_ok=True)
    with open(OUT_INDEX, "w") as f:
        json.dump({"embed_model": EMBED_MODEL, "dims": EMBED_DIMS,
                   "anchor_date": anchor.isoformat(),
                   "songs": records, "artists": artists}, f, ensure_ascii=False)
    size = os.path.getsize(OUT_INDEX) / 1e6
    print(f"Compiled {len(records)} songs + {len(artists)} artists -> {OUT_INDEX} ({size:.1f}MB)")


def build_artists(records):
    """Artist-level stats for structured intents (popular / collaborators /
    near-me / new-artists): profile fields from the User sheet, activity dates
    from the FULL Songs sheet (no API cost), aggregated audio traits from the
    enriched subset."""
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    def sheet(name):
        rows = wb[name].iter_rows(values_only=True)
        hdr = next(rows)
        return [dict(zip(hdr, r)) for r in rows if any(c is not None for c in r)]

    users = sheet("User")
    all_songs = sheet("Songs")
    dates_by_artist = defaultdict(list)
    for s in all_songs:
        if s["UPLOAD_DATE"]:
            dates_by_artist[s["USER_ID"]].append(str(s["UPLOAD_DATE"]))
    enriched_by_artist = defaultdict(list)
    for r in records:
        enriched_by_artist[r["artist_id"]].append(r)

    artists = []
    for u in users:
        aid = u["USER_ID"]
        dates = sorted(dates_by_artist.get(aid, []))
        er = enriched_by_artist.get(aid, [])
        moods = Counter(m for r in er for m in r["analysis"]["moods"])
        instruments = Counter(i for r in er for i in r["analysis"]["instruments"])
        name = str(u["ARTIST_NAME"])
        artists.append({
            "artist_id": aid,
            "name": name,
            "region": u["REGION"],
            "rating": u["RATING_SCORE"] or 0,
            "rank": u["USER_RANK"] or 999,
            "follow_count": u["FOLLOW_COUNT"] or 0,
            "follower_count": u["FOLLOWER_COUNT"] or 0,
            "total_playback": u["PLAYBACK_COUNT"] or 0,
            "song_count": u["SONG_COUNT"] or 0,
            "genres": [g.strip() for g in str(u["GENRES"] or "").split(",") if g.strip()],
            "enriched_songs": len(er),
            "first_seen": dates[0] if dates else None,
            "last_seen": dates[-1] if dates else None,
            "export_capped": bool(u["SONG_COUNT"] and u["SONG_COUNT"] > len(dates)),
            "top_moods": [m for m, _ in moods.most_common(5)],
            "top_instruments": [i for i, _ in instruments.most_common(5)],
            "is_house": name in HOUSE_ACCOUNTS,
        })
    return artists


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=150)
    ap.add_argument("--workers", type=int, default=4)
    ap.add_argument("--compile", action="store_true", help="only rebuild web index")
    args = ap.parse_args()
    if args.compile:
        compile_index()
    else:
        run(args.limit, args.workers)
        compile_index()
